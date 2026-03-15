"""Import player data from Mr. Cheatsheet Excel file into players.json."""

import json
import sys
from pathlib import Path

import openpyxl


# Column mappings for Mr. Cheatsheet's Special Blend projections
# Hitting: columns DR(122) through EH(138) in Projection-H (0-indexed: 121-137)
HITTING_BLEND_COLS = {
    "PA": 121,   # DR
    "AB": 122,   # DS
    "R": 123,    # DT
    "H": 124,    # DU
    "1B": 125,   # DV
    "2B": 126,   # DW
    "3B": 127,   # DX
    "HR": 128,   # DY
    "RBI": 129,  # DZ
    "SB": 130,   # EA
    "CS": 131,   # EB
    "BB": 132,   # EC
    "K": 133,    # ED
    "AVG": 134,  # EE
    "OBP": 135,  # EF
    "SLG": 136,  # EG
    "OPS": 137,  # EH
}

# Pitching: columns DD(108) through DR(122) in Projection-P (0-indexed: 107-121)
PITCHING_BLEND_COLS = {
    "W": 107,   # DD
    "L": 108,   # DE
    "IP": 109,  # DF
    "H": 110,   # DG (Hits Allowed)
    "BB": 111,  # DH (Walks)
    "K": 112,   # DI
    "ER": 113,  # DJ
    "SV": 114,  # DK
    "ERA": 115, # DL
    "WHIP": 116,# DM
    "K/9": 117, # DN
    "QS": 118,  # DO
    "GS": 119,  # DP
    "G": 120,   # DQ
    "HR": 121,  # DR
}

# Positional eligibility columns in Pre sheet (0-indexed: 22-33)
POS_COLS = {
    22: "C", 23: "1B", 24: "2B", 25: "3B", 26: "SS",
    27: "LF", 28: "CF", 29: "RF", 30: "OF", 31: "DH",
    32: "SP", 33: "RP",
}


def safe_float(val, default=0.0):
    if val is None or val == "" or val == "#N/A":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def import_players(excel_path: str, output_path: str) -> None:
    path = Path(excel_path)
    if not path.exists():
        print(f"Error: {path} not found")
        sys.exit(1)

    print(f"Loading {path.name}...")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True, keep_vba=True)

    # Build player index from Core-Data
    core = wb["Core-Data"]
    players = {}
    for row in core.iter_rows(min_row=2):
        name = row[0].value  # A: PLAYERNAME
        if not name or name == "-":
            continue
        player_id = row[1].value or ""  # B: IDPLAYER
        main_pos = row[2].value or ""   # C: MC Pos
        team = row[4].value or ""       # E: TEAM
        age = row[5].value or 0         # F: Age

        # ADP: use FantasyPros Avg (col M, idx 12) with ESPN (col L, idx 11) as fallback
        adp = safe_float(row[12].value, 999)
        if adp == 999:
            adp = safe_float(row[11].value, 999)

        players[name] = {
            "name": name,
            "player_id": str(player_id),
            "team": team,
            "age": int(safe_float(age, 0)),
            "main_position": main_pos,
            "positions": [main_pos] if main_pos else [],
            "adp": round(adp, 1),
            "hitting_projections": {},
            "pitching_projections": {},
        }

    print(f"  Found {len(players)} players in Core-Data")

    # Add positional eligibility from Pre sheet
    pre = wb["Pre"]
    for row in pre.iter_rows(min_row=2):
        name = row[0].value
        if not name or name == "-" or name not in players:
            continue
        positions = set()
        for col_idx, pos_name in POS_COLS.items():
            val = row[col_idx].value
            if val is not None and safe_float(val, 0) >= 1:
                positions.add(pos_name)
        if positions:
            players[name]["positions"] = sorted(positions)
        elif players[name]["main_position"]:
            players[name]["positions"] = [players[name]["main_position"]]

    # Add hitting projections from Projection-H
    proj_h = wb["Projection-H"]
    hit_count = 0
    for row in proj_h.iter_rows(min_row=2):
        name = row[0].value  # A: PLAYERNAME
        if not name or name not in players:
            continue
        projections = {}
        for stat, col_idx in HITTING_BLEND_COLS.items():
            val = row[col_idx].value
            if val is not None and val != "#N/A":
                projections[stat] = round(safe_float(val), 4)
        if projections:
            players[name]["hitting_projections"] = projections
            hit_count += 1

    print(f"  Loaded hitting projections for {hit_count} players")

    # Add pitching projections from Projection-P
    proj_p = wb["Projection-P"]
    pitch_count = 0
    for row in proj_p.iter_rows(min_row=2):
        name = row[0].value  # A: PLAYERNAME
        if not name or name not in players:
            continue
        projections = {}
        for stat, col_idx in PITCHING_BLEND_COLS.items():
            val = row[col_idx].value
            if val is not None and val != "#N/A":
                projections[stat] = round(safe_float(val), 4)
        if projections:
            players[name]["pitching_projections"] = projections
            pitch_count += 1

    print(f"  Loaded pitching projections for {pitch_count} players")

    wb.close()

    # Write output
    player_list = sorted(players.values(), key=lambda p: p["adp"])
    out = Path(output_path)
    with open(out, "w") as f:
        json.dump(player_list, f, indent=2)

    print(f"\nWrote {len(player_list)} players to {out}")
    print(f"Top 10 by ADP:")
    for p in player_list[:10]:
        pos = "/".join(p["positions"])
        print(f"  {p['adp']:>6.1f}  {p['name']:<25} {pos:<12} {p['team']}")


if __name__ == "__main__":
    excel_file = sys.argv[1] if len(sys.argv) > 1 else "2026_Roto_Draft_Cheatsheet_v1.00.xlsm"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "data/players.json"
    import_players(excel_file, output_file)
